# app.py
from flask import Flask, render_template, request, send_file
from parser import parse_kivano
import pandas as pd
import os
import requests
from openpyxl import Workbook
from openpyxl.drawing.image import Image as XLImage
from openpyxl.utils import get_column_letter
from io import BytesIO
from PIL import Image as PILImage

app = Flask(__name__)
EXCEL_PATH = "static/results.xlsx"

def save_to_excel_with_images(products, output_path):
    # Создаём новую книгу
    wb = Workbook()
    ws = wb.active
    ws.title = "Товары"

    # Заголовки
    headers = ["Фото", "Название", "Цена (сом)", "Ссылка"]
    ws.append(headers)

    # Устанавливаем ширину столбцов
    ws.column_dimensions['A'].width = 15  # фото
    ws.column_dimensions['B'].width = 50  # название
    ws.column_dimensions['C'].width = 15  # цена
    ws.column_dimensions['D'].width = 60  # ссылка

    row_num = 2  # первая строка — заголовки

    for product in products:
        # Записываем текстовые данные
        ws.cell(row=row_num, column=2, value=product["Название"])
        ws.cell(row=row_num, column=3, value=product["Цена (сом)"])
        ws.cell(row=row_num, column=4, value=product["Ссылка"])

        # Обработка изображения
        img_url = product.get("Изображение")
        if img_url:
            try:
                response = requests.get(img_url, timeout=5)
                if response.status_code == 200:
                    # Открываем изображение через PIL
                    img = PILImage.open(BytesIO(response.content))
                    # Масштабируем до 100x100 (сохраняя пропорции)
                    img.thumbnail((100, 100))
                    # Сохраняем во временный буфер
                    img_buffer = BytesIO()
                    img.save(img_buffer, format="PNG")
                    img_buffer.seek(0)

                    # Создаём изображение для openpyxl
                    xl_img = XLImage(img_buffer)
                    xl_img.width = 100
                    xl_img.height = 100

                    # Вставляем в ячейку A{row}
                    ws.add_image(xl_img, f"A{row_num}")

                    # Подгоняем высоту строки
                    ws.row_dimensions[row_num].height = 80
            except Exception as e:
                print(f"Не удалось загрузить фото: {e}")

        row_num += 1

    # Сохраняем файл
    wb.save(output_path)

@app.route('/', methods=['GET', 'POST'])
def index():
    if request.method == 'POST':
        query = request.form.get('query', '').strip()
        if not query:
            return render_template('index.html', error="Введите запрос!")

        products = parse_kivano(query)
        summary = f"Найдено {len(products)} товаров по запросу '{query}'."

        if products:
            os.makedirs("static", exist_ok=True)
            save_to_excel_with_images(products, EXCEL_PATH)

        return render_template('index.html', summary=summary, products=products[:5], query=query)

    return render_template('index.html')

@app.route('/download')
def download():
    if os.path.exists(EXCEL_PATH):
        return send_file(EXCEL_PATH, as_attachment=True, download_name="товары_с_фото.xlsx")
    return "Файл не найден", 404

if __name__ == '__main__':
    app.run(debug=True)